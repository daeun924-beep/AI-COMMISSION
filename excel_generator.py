import io
from openpyxl import load_workbook
from openpyxl.styles import Font

def set_value_only(ws, row, col, value, is_number=False):
    """지정한 좌표(row, col)에 유령 데이터를 무시하고 값을 곧바로 강제 주입하는 함수"""
    cell = ws.cell(row=row, column=col)
    
    # None 값이 들어오면 빈 문자열로 대체하여 에러 방지
    cell.value = "" if value is None else value
    
    if cell.font:
        cell.font = Font(
            name=cell.font.name or "맑은 고딕", 
            size=cell.font.size or 10, 
            bold=cell.font.bold, 
            color=cell.font.color
        )
        
    if is_number and isinstance(value, (int, float)):
        cell.number_format = '#,##0'

def generate_quotation_excel(template_path, detail_rows, calc_results, quantity, discount_rate, client_info):
    """
    요금합계(E39)를 총금액 - 부가세 공식으로 안전하게 연산하여 기입하는 최종 마스터 함수
    """
    wb = load_workbook(template_path)
    ws = wb.active
    
    # [1] 상단 의뢰자 정보 영역 -> 무조건 E열(5번째 열)에 강제 기입
    set_value_only(ws, 4, 5, client_info.get("company", ""))  # E4 셀 (업체명)
    set_value_only(ws, 5, 5, client_info.get("manager", ""))  # E5 셀 (담당자)
    set_value_only(ws, 6, 5, client_info.get("sample", ""))   # E6 셀 (시료명)

    # [2] 중앙 시험 항목 영역 기입 (A~E열 정위치)
    for idx, row in enumerate(detail_rows[:28]):
        r = 9 + idx
        set_value_only(ws, r, 1, row["name"])              # A열 (시험내역)
        set_value_only(ws, r, 2, row["price"], is_number=True) # B열 (수수료)
        set_value_only(ws, r, 3, quantity, is_number=True)  # C열 (시료수)
        set_value_only(ws, r, 4, f"{discount_rate}%")      # D열 (할인율)
        
        pure_price = row["price"] * quantity
        discounted_price = int(pure_price * (1 - discount_rate / 100))
        set_value_only(ws, r, 5, discounted_price, is_number=True) # E열 (금액)

    # [3] 하단 요약 영역 고정 좌표 (E열 = 5번째 열)
    # 딕셔너리에서 안전하게 최종 금액과 부가세를 먼저 가져옵니다.
    total_amt = calc_results.get("total", 0)
    vat_amt = calc_results.get("vat", 0)
    
    # 💡 유저님이 말씀하신 규칙 적용: 요금합계 = 총금액 - 부가세
    subtotal_amt = total_amt - vat_amt

    set_value_only(ws, 37, 5, calc_results.get("extra", 0), is_number=True)       # E37 : 추가요금
    set_value_only(ws, 38, 5, calc_results.get("etc", 0), is_number=True)         # E38 : 기본/기타비용
    set_value_only(ws, 39, 5, subtotal_amt, is_number=True)                       # E39 : 요금합계 (보정 완료)
    set_value_only(ws, 41, 5, vat_amt, is_number=True)                            # E41 : 부가세
    set_value_only(ws, 42, 5, total_amt, is_number=True)                          # E42 : 총금액

    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    
    return output_stream.getvalue()